import os
import io
import chardet
from datetime import datetime
from fastapi import APIRouter, HTTPException, Header, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from pydantic import BaseModel
from app.database import db
from app.auth_utils import create_token, verify_teacher_token
from app.sync_exams import sync_exams
from pypinyin import lazy_pinyin, Style
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

TEACHER_PASSWORD = os.environ.get("TEACHER_PASSWORD", "admin123")


def _require_teacher(authorization: Optional[str]):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="请先登录")
    token = authorization.removeprefix("Bearer ").strip()
    try:
        verify_teacher_token(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))


def _name_to_pinyin(name: str):
    full = "".join(lazy_pinyin(name, style=Style.NORMAL))
    abbr = "".join(lazy_pinyin(name, style=Style.FIRST_LETTER))
    return full.lower(), abbr.lower()


class LoginRequest(BaseModel):
    password: str


@router.post("/api/teacher/login")
def teacher_login(req: LoginRequest):
    if req.password != TEACHER_PASSWORD:
        raise HTTPException(status_code=401, detail="密码错误")
    token = create_token({"role": "teacher"}, expires_hours=8)
    return {"token": token}


@router.post("/api/teacher/exam_students")
async def upload_students(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None)
):
    """上传学生名单 CSV（UTF-8 或 GBK 均支持）"""
    _require_teacher(authorization)

    raw = await file.read()
    encoding = chardet.detect(raw)["encoding"] or "utf-8"
    text = raw.decode(encoding, errors="replace")

    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        raise HTTPException(status_code=422, detail="文件为空")

    # 跳过表头
    data_lines = lines[1:] if lines[0].startswith("姓名") or lines[0].startswith("name") else lines

    records = []
    for line in data_lines:
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 2:
            continue
        name = parts[0]
        student_id = parts[1]
        class_name = parts[2] if len(parts) > 2 else ""
        if not name or not student_id:
            continue
        pinyin, abbr = _name_to_pinyin(name)
        records.append((name, student_id, class_name, pinyin, abbr))

    if not records:
        raise HTTPException(status_code=422, detail="CSV 中没有有效数据行")

    with db() as conn:
        cur = conn.cursor()
        cur.executemany("""
            INSERT INTO exam_students (name, student_id, class_name, pinyin, pinyin_abbr)
            VALUES (%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                name=VALUES(name),
                class_name=VALUES(class_name),
                pinyin=VALUES(pinyin),
                pinyin_abbr=VALUES(pinyin_abbr)
        """, records)
        cur.close()

    return {"count": len(records)}


class AddStudentRequest(BaseModel):
    name: str
    student_id: str
    class_name: str = ""


@router.post("/api/teacher/exam_students/add")
def add_student(req: AddStudentRequest, authorization: Optional[str] = Header(None)):
    """添加单个学生"""
    _require_teacher(authorization)
    req.name = req.name.strip()
    req.student_id = req.student_id.strip()
    if not req.name or not req.student_id:
        raise HTTPException(status_code=422, detail="姓名和学号不能为空")
    pinyin, abbr = _name_to_pinyin(req.name)
    with db() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM exam_students WHERE student_id=%s", (req.student_id,)
        )
        existing = cur.fetchone()
        if existing:
            raise HTTPException(status_code=409, detail=f"学号 {req.student_id} 已存在（{existing['name']}）")
        cur.execute(
            "INSERT INTO exam_students (name, student_id, class_name, pinyin, pinyin_abbr) VALUES (%s,%s,%s,%s,%s)",
            (req.name, req.student_id, req.class_name.strip(), pinyin, abbr)
        )
        cur.close()
    return {"ok": True}


@router.delete("/api/teacher/exam_students/{student_id}")
def delete_student(student_id: str, authorization: Optional[str] = Header(None)):
    """删除单个学生（同时删除其成绩）"""
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM exam_students WHERE student_id=%s", (student_id,)
        )
        student = cur.fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="学号不存在")
        cur.execute("DELETE FROM exam_scores WHERE student_id=%s", (student_id,))
        cur.execute("DELETE FROM exam_students WHERE student_id=%s", (student_id,))
        cur.close()
    return {"ok": True, "deleted": student["name"]}


@router.delete("/api/teacher/exam_students")
def clear_all_students(authorization: Optional[str] = Header(None)):
    """清空全部学生名单（同时清空所有成绩）"""
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS cnt FROM exam_students")
        count = cur.fetchone()["cnt"]
        cur.execute("DELETE FROM exam_scores")
        cur.execute("DELETE FROM exam_students")
        cur.close()
    return {"ok": True, "deleted_count": count}


@router.post("/api/teacher/reset")
def new_semester_reset(authorization: Optional[str] = Header(None)):
    """新学期重置：清空所有学生名单 + 所有成绩"""
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS cnt FROM exam_students")
        exam_students = cur.fetchone()["cnt"]
        cur.execute("SELECT COUNT(*) AS cnt FROM exam_scores")
        exam_scores = cur.fetchone()["cnt"]
        cur.execute("DELETE FROM exam_scores")
        cur.execute("DELETE FROM exam_students")
        cur.close()
    return {"ok": True, "deleted_students": exam_students, "deleted_scores": exam_scores}


@router.get("/api/teacher/exam_students/list")
def list_students(authorization: Optional[str] = Header(None)):
    """获取全部学生名单"""
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT name, student_id, class_name FROM exam_students ORDER BY class_name, name"
        )
        rows = cur.fetchall()
        cur.close()
    return [dict(r) for r in rows]


@router.get("/api/teacher/exam_list")
def list_exams(authorization: Optional[str] = Header(None)):
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id, title, is_active FROM exam_list ORDER BY id")
        exam_list = cur.fetchall()
        cur.close()

    # 懒加载同步：若数据库中还没有考试记录，尝试立即扫描文档目录
    if not exam_list:
        print("[list_exams] 考试表为空，触发懒加载同步…")
        sync_exams()
        with db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT id, title, is_active FROM exam_list ORDER BY id")
            exam_list = cur.fetchall()
            cur.close()

    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS cnt FROM exam_students")
        total_students = cur.fetchone()["cnt"]
        result = []
        for e in exam_list:
            cur.execute(
                "SELECT COUNT(*) AS cnt FROM exam_scores WHERE exam_id=%s", (e["id"],)
            )
            submitted = cur.fetchone()["cnt"]
            cur.execute(
                "SELECT AVG(score) FROM exam_scores WHERE exam_id=%s", (e["id"],)
            )
            avg = cur.fetchone()["AVG(score)"]
            result.append({
                "id": e["id"], "title": e["title"], "is_active": e["is_active"],
                "submitted": submitted, "total_students": total_students,
                "avg_score": round(float(avg), 1) if avg else None,
            })
        cur.close()
    return result


class ExamCreate(BaseModel):
    id: str
    title: str


@router.post("/api/teacher/exam_list")
def create_exam(req: ExamCreate, authorization: Optional[str] = Header(None)):
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute(
            "REPLACE INTO exam_list (id, title) VALUES (%s,%s)",
            (req.id, req.title)
        )
        cur.close()
    return {"ok": True}


class ExamUpdate(BaseModel):
    is_active: int


@router.put("/api/teacher/exam_list/{exam_id}")
def update_exam(exam_id: str, req: ExamUpdate, authorization: Optional[str] = Header(None)):
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE exam_list SET is_active=%s WHERE id=%s", (req.is_active, exam_id))
        cur.close()
    return {"ok": True}


@router.get("/api/teacher/exam_scores")
def get_scores(exam_id: str = Query(...), authorization: Optional[str] = Header(None)):
    """获取某次考试的所有学生成绩（含未提交）"""
    _require_teacher(authorization)
    import json as _json

    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT title FROM exam_list WHERE id=%s", (exam_id,))
        exam = cur.fetchone()
        if not exam:
            raise HTTPException(status_code=404, detail="考试不存在")

        cur.execute(
            "SELECT name, student_id, class_name FROM exam_students ORDER BY student_id"
        )
        exam_students = cur.fetchall()
        cur.execute(
            "SELECT student_id, score, total, submitted_at, allow_resubmit FROM exam_scores WHERE exam_id=%s",
            (exam_id,)
        )
        scores_rows = cur.fetchall()
        scores_map = {r["student_id"]: dict(r) for r in scores_rows}

        # 查询该考试所有评语
        feedbacks_map = {}
        cur.execute(
            "SELECT student_id, id, teacher_comment, deduction_items, suggestions, created_at "
            "FROM exam_feedbacks WHERE exam_id=%s ORDER BY student_id, created_at DESC",
            (exam_id,)
        )
        all_fbs = cur.fetchall()
        for fb in all_fbs:
            sid = fb["student_id"]
            if sid not in feedbacks_map:
                feedbacks_map[sid] = []
            feedbacks_map[sid].append({
                "id": fb["id"],
                "teacher_comment": fb["teacher_comment"],
                "deduction_items": _json.loads(fb["deduction_items"]) if isinstance(fb["deduction_items"], str) else (fb["deduction_items"] or []),
                "suggestions": _json.loads(fb["suggestions"]) if isinstance(fb["suggestions"], str) else (fb["suggestions"] or []),
                "created_at": str(fb["created_at"]) if fb["created_at"] else None,
            })
        cur.close()

    result = []
    for s in exam_students:
        sc = scores_map.get(s["student_id"])
        result.append({
            "name": s["name"],
            "student_id": s["student_id"],
            "class_name": s["class_name"],
            "score": sc["score"] if sc else None,
            "total": sc["total"] if sc else None,
            "submitted_at": str(sc["submitted_at"]) if sc and sc["submitted_at"] else None,
            "allow_resubmit": bool(sc["allow_resubmit"]) if sc else False,
            "exam_feedbacks": feedbacks_map.get(s["student_id"], []),
        })

    return {"exam_title": exam["title"], "rows": result}


@router.get("/api/teacher/exam_scores/export")
def export_scores(exam_id: str = Query(...), authorization: Optional[str] = Header(None)):
    """导出成绩 Excel"""
    _require_teacher(authorization)

    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT title FROM exam_list WHERE id=%s", (exam_id,))
        exam = cur.fetchone()
        if not exam:
            raise HTTPException(status_code=404, detail="考试不存在")

        cur.execute(
            "SELECT name, student_id, class_name FROM exam_students ORDER BY student_id"
        )
        exam_students = cur.fetchall()
        cur.execute(
            "SELECT student_id, score, total, submitted_at FROM exam_scores WHERE exam_id=%s",
            (exam_id,)
        )
        scores_rows = cur.fetchall()
        scores_map = {r["student_id"]: dict(r) for r in scores_rows}
        cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = exam["title"][:31]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["姓名", "学号", "班级", "得分", "满分", "提交时间"]
    col_widths = [12, 14, 20, 8, 8, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for row_idx, s in enumerate(exam_students, 2):
        sc = scores_map.get(s["student_id"])
        ws.cell(row=row_idx, column=1, value=s["name"])
        ws.cell(row=row_idx, column=2, value=s["student_id"])
        ws.cell(row=row_idx, column=3, value=s["class_name"])
        ws.cell(row=row_idx, column=4, value=sc["score"] if sc else "")
        ws.cell(row=row_idx, column=5, value=sc["total"] if sc else "")
        ws.cell(row=row_idx, column=6, value=str(sc["submitted_at"]) if sc and sc["submitted_at"] else "")
        if not sc:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).font = Font(color="999999")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"成绩单_{exam['title']}_{date_str}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ───── 评语与反馈管理 ─────

class FeedbackCreate(BaseModel):
    student_id: str
    exam_id: str
    teacher_comment: str = ""
    deduction_items: list = []
    suggestions: list = []
    allow_resubmit: bool = False


class FeedbackUpdate(BaseModel):
    teacher_comment: Optional[str] = None
    deduction_items: Optional[list] = None
    suggestions: Optional[list] = None
    allow_resubmit: Optional[bool] = None


@router.post("/api/teacher/exam_feedbacks")
def create_feedback(req: FeedbackCreate, authorization: Optional[str] = Header(None)):
    """为某学生的某次考试添加评语、扣分项和改进建议"""
    _require_teacher(authorization)
    import json as _json

    with db() as conn:
        cur = conn.cursor()
        # 检查学生是否存在
        cur.execute(
            "SELECT name FROM exam_students WHERE student_id=%s", (req.student_id,)
        )
        student = cur.fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="学号不存在")

        # 检查考试是否存在
        cur.execute(
            "SELECT title FROM exam_list WHERE id=%s", (req.exam_id,)
        )
        exam = cur.fetchone()
        if not exam:
            raise HTTPException(status_code=404, detail="考试不存在")

        deduction_json = _json.dumps(req.deduction_items, ensure_ascii=False)
        suggestions_json = _json.dumps(req.suggestions, ensure_ascii=False)

        cur.execute(
            "INSERT INTO exam_feedbacks (student_id, exam_id, teacher_comment, deduction_items, suggestions) "
            "VALUES (%s,%s,%s,%s,%s)",
            (req.student_id, req.exam_id, req.teacher_comment, deduction_json, suggestions_json)
        )
        feedback_id = cur.lastrowid

        # 如果勾选了允许二次提交，更新 exam_scores 表
        if req.allow_resubmit:
            cur.execute(
                "UPDATE exam_scores SET allow_resubmit=1 WHERE student_id=%s AND exam_id=%s",
                (req.student_id, req.exam_id)
            )
        cur.close()

    return {
        "ok": True,
        "id": feedback_id,
        "student_id": req.student_id,
        "exam_id": req.exam_id,
        "allow_resubmit": req.allow_resubmit,
    }


@router.get("/api/teacher/exam_feedbacks")
def list_feedbacks(
    exam_id: str = Query(...),
    student_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None)
):
    """教师查看评语列表，可按考试和学号筛选"""
    _require_teacher(authorization)
    import json as _json

    with db() as conn:
        cur = conn.cursor()
        if student_id:
            cur.execute(
                "SELECT f.id, f.student_id, f.exam_id, f.teacher_comment, f.deduction_items, "
                "f.suggestions, f.created_at, f.updated_at, s.allow_resubmit "
                "FROM exam_feedbacks f LEFT JOIN exam_scores s ON f.student_id=s.student_id AND f.exam_id=s.exam_id "
                "WHERE f.exam_id=%s AND f.student_id=%s ORDER BY f.created_at DESC",
                (exam_id, student_id)
            )
        else:
            cur.execute(
                "SELECT f.id, f.student_id, f.exam_id, f.teacher_comment, f.deduction_items, "
                "f.suggestions, f.created_at, f.updated_at, s.allow_resubmit "
                "FROM exam_feedbacks f LEFT JOIN exam_scores s ON f.student_id=s.student_id AND f.exam_id=s.exam_id "
                "WHERE f.exam_id=%s ORDER BY f.student_id, f.created_at DESC",
                (exam_id,)
            )
        rows = cur.fetchall()
        cur.close()

    result = []
    for r in rows:
        deduction = r["deduction_items"]
        suggestion = r["suggestions"]
        if isinstance(deduction, str):
            deduction = _json.loads(deduction) if deduction else []
        if isinstance(suggestion, str):
            suggestion = _json.loads(suggestion) if suggestion else []
        result.append({
            "id": r["id"],
            "student_id": r["student_id"],
            "exam_id": r["exam_id"],
            "teacher_comment": r["teacher_comment"],
            "deduction_items": deduction or [],
            "suggestions": suggestion or [],
            "allow_resubmit": bool(r["allow_resubmit"]) if r["allow_resubmit"] is not None else False,
            "created_at": str(r["created_at"]) if r["created_at"] else None,
            "updated_at": str(r["updated_at"]) if r["updated_at"] else None,
        })

    return {"exam_feedbacks": result}


@router.put("/api/teacher/exam_feedbacks/{feedback_id}")
def update_feedback(feedback_id: int, req: FeedbackUpdate, authorization: Optional[str] = Header(None)):
    """更新评语内容"""
    _require_teacher(authorization)
    import json as _json

    with db() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT student_id, exam_id FROM exam_feedbacks WHERE id=%s", (feedback_id,)
        )
        fb = cur.fetchone()
        if not fb:
            raise HTTPException(status_code=404, detail="评语不存在")

        updates = []
        params = []

        if req.teacher_comment is not None:
            updates.append("teacher_comment=%s")
            params.append(req.teacher_comment)
        if req.deduction_items is not None:
            updates.append("deduction_items=%s")
            params.append(_json.dumps(req.deduction_items, ensure_ascii=False))
        if req.suggestions is not None:
            updates.append("suggestions=%s")
            params.append(_json.dumps(req.suggestions, ensure_ascii=False))

        if updates:
            params.append(feedback_id)
            cur.execute(
                f"UPDATE exam_feedbacks SET {', '.join(updates)} WHERE id=%s",
                params
            )

        # 更新二次提交状态
        if req.allow_resubmit is not None:
            cur.execute(
                "UPDATE exam_scores SET allow_resubmit=%s WHERE student_id=%s AND exam_id=%s",
                (1 if req.allow_resubmit else 0, fb["student_id"], fb["exam_id"])
            )
        cur.close()

    return {"ok": True}


@router.delete("/api/teacher/exam_feedbacks/{feedback_id}")
def delete_feedback(feedback_id: int, authorization: Optional[str] = Header(None)):
    """删除评语"""
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM exam_feedbacks WHERE id=%s", (feedback_id,))
        fb = cur.fetchone()
        if not fb:
            raise HTTPException(status_code=404, detail="评语不存在")
        cur.execute("DELETE FROM exam_feedbacks WHERE id=%s", (feedback_id,))
        cur.close()
    return {"ok": True}


@router.put("/api/teacher/exam_scores/{student_id}/{exam_id}/resubmit")
def toggle_resubmit(student_id: str, exam_id: str, authorization: Optional[str] = Header(None)):
    """开启/关闭某学生的二次提交权限"""
    _require_teacher(authorization)
    with db() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT id, allow_resubmit FROM exam_scores WHERE student_id=%s AND exam_id=%s",
            (student_id, exam_id)
        )
        score = cur.fetchone()
        if not score:
            raise HTTPException(status_code=404, detail="该学生尚未提交本次考试")

        new_val = 0 if score["allow_resubmit"] else 1
        cur.execute(
            "UPDATE exam_scores SET allow_resubmit=%s WHERE student_id=%s AND exam_id=%s",
            (new_val, student_id, exam_id)
        )
        cur.close()

    return {"ok": True, "student_id": student_id, "exam_id": exam_id, "allow_resubmit": bool(new_val)}
