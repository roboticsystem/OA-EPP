import io
import os
import zipfile
from datetime import datetime
from typing import Optional, List
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from app.models.report_models import ReportData


class ExportService:

    @staticmethod
    def generate_pdf(report: ReportData) -> bytes:
        if not HAS_REPORTLAB:
            raise Exception("PDF生成库未安装，请运行: pip install reportlab")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor('#1565c0')
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor('#1565c0'),
            borderColor=colors.HexColor('#1565c0'),
            borderWidth=0,
            borderPadding=5
        )

        normal_style = styles['Normal']
        normal_style.fontSize = 10

        elements = []

        elements.append(Paragraph(report.course_name, title_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("学生开发日志报告", ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor('#333333')
        )))

        info_data = [
            ['姓名', report.student_info.name],
            ['学号', report.student_info.student_id],
            ['班级', report.student_info.class_name or '未分班'],
            ['学期', report.semester],
            ['生成时间', report.generated_at]
        ]
        info_table = Table(info_data, colWidths=[3*cm, 10*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("一、分支记录", heading_style))
        if report.branches:
            branch_data = [['分支名', '保护', '最后提交']]
            for b in report.branches[:10]:
                branch_data.append([
                    b.name,
                    '是' if b.protected else '否',
                    b.last_commit_date[:10] if b.last_commit_date else '-'
                ])
            branch_table = Table(branch_data, colWidths=[5*cm, 2*cm, 6*cm])
            branch_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(branch_table)
            elements.append(Paragraph(f"共 {len(report.branches)} 个分支", ParagraphStyle('Small', fontSize=8, textColor=colors.gray)))
        else:
            elements.append(Paragraph("暂无分支记录", normal_style))

        elements.append(Spacer(1, 15))
        elements.append(Paragraph("二、提交历史", heading_style))
        if report.commits:
            commit_data = [['SHA', '提交信息', '作者', '日期']]
            for c in report.commits[:15]:
                commit_data.append([
                    c.sha[:7],
                    c.message[:40] + ('...' if len(c.message) > 40 else ''),
                    c.author_name,
                    c.date[:10] if c.date else '-'
                ])
            commit_table = Table(commit_data, colWidths=[2*cm, 6*cm, 3*cm, 2.5*cm])
            commit_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(commit_table)
            elements.append(Paragraph(f"共 {len(report.commits)} 次提交", ParagraphStyle('Small', fontSize=8, textColor=colors.gray)))
        else:
            elements.append(Paragraph("暂无提交记录", normal_style))

        elements.append(PageBreak())
        elements.append(Paragraph("三、代码质量", heading_style))
        cq = report.code_quality
        elements.append(Paragraph(f"总代码行数: {cq.total_lines:,} 行", normal_style))
        if cq.languages:
            lang_data = [['语言', '占比']]
            for lang in cq.top_languages[:5]:
                lang_data.append([lang['name'], f"{lang['percentage']}%"])
            lang_table = Table(lang_data, colWidths=[5*cm, 5*cm])
            lang_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f5e9')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ]))
            elements.append(lang_table)

        elements.append(Spacer(1, 15))
        elements.append(Paragraph("四、Pull Request 情况", heading_style))
        if report.pull_requests:
            pr_data = [['#', '标题', '状态', '创建时间']]
            for pr in report.pull_requests[:10]:
                state_text = '已合并' if pr.merged_at else pr.state
                pr_data.append([
                    str(pr.number),
                    pr.title[:35] + ('...' if len(pr.title) > 35 else ''),
                    state_text,
                    pr.created_at[:10] if pr.created_at else '-'
                ])
            pr_table = Table(pr_data, colWidths=[1.5*cm, 7*cm, 2.5*cm, 2.5*cm])
            pr_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fff3e0')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(pr_table)
        else:
            elements.append(Paragraph("暂无PR记录", normal_style))

        elements.append(Spacer(1, 15))
        elements.append(Paragraph("五、PR 分析统计", heading_style))
        pra = report.pr_analysis
        pr_stats_data = [
            ['总PR数', str(pra.total_prs)],
            ['已合并', str(pra.merged_prs)],
            ['合并率', f"{pra.merge_rate}%"],
            ['总新增行', f"+{pra.total_additions}"],
            ['总删除行', f"-{pra.total_deletions}"]
        ]
        pr_stats_table = Table(pr_stats_data, colWidths=[4*cm, 9*cm])
        pr_stats_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(pr_stats_table)

        elements.append(PageBreak())
        elements.append(Paragraph("六、教师评语", heading_style))
        if report.teacher_comments:
            for tc in report.teacher_comments:
                comment_date = tc.created_at[:10] if tc.created_at else ''
                elements.append(Paragraph(f"[{comment_date}] {tc.comment}", normal_style))
                elements.append(Spacer(1, 8))
        else:
            elements.append(Paragraph("暂无评语", normal_style))

        elements.append(Spacer(1, 15))
        elements.append(Paragraph("七、在线考试成绩", heading_style))
        if report.exam_scores:
            exam_data = [['考试名称', '得分', '总分', '提交时间']]
            for e in report.exam_scores:
                exam_data.append([
                    e.exam_title,
                    str(e.score),
                    str(e.total),
                    e.submitted_at[:10] if e.submitted_at else '-'
                ])
            exam_table = Table(exam_data, colWidths=[5*cm, 2.5*cm, 2.5*cm, 3.5*cm])
            exam_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(exam_table)
        else:
            elements.append(Paragraph("暂无考试记录", normal_style))

        elements.append(Spacer(1, 15))
        elements.append(Paragraph("八、考勤记录", heading_style))
        att = report.attendance_summary
        att_data = [
            ['总天数', str(att.total_days)],
            ['出勤', f"{att.present_days}天"],
            ['缺勤', f"{att.absent_days}天"],
            ['迟到', f"{att.late_days}天"],
            ['请假', f"{att.leave_days}天"],
            ['出勤率', f"{att.attendance_rate}%"]
        ]
        att_table = Table(att_data, colWidths=[4*cm, 9*cm])
        att_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 5), (-1, 5), colors.HexColor('#2e7d32')),
        ]))
        elements.append(att_table)

        elements.append(Spacer(1, 15))
        elements.append(Paragraph("九、课程总评", heading_style))
        cs = report.course_summary
        summary_data = [
            ['完成考试', f"{cs.completed_exams}/{cs.total_exams}次"],
            ['平均分', f"{cs.avg_score}分"],
            ['总分', f"{cs.total_score}分"],
            ['出勤率', f"{att.attendance_rate}%"]
        ]
        summary_table = Table(summary_data, colWidths=[4*cm, 9*cm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1565c0')),
        ]))
        elements.append(summary_table)

        elements.append(Spacer(1, 30))
        elements.append(Paragraph("— 报告由系统自动生成 —", ParagraphStyle('Footer', fontSize=8, textColor=colors.gray, alignment=1)))
        elements.append(Paragraph("成都信息工程大学 软件工程学院", ParagraphStyle('Footer', fontSize=9, textColor=colors.gray, alignment=1)))

        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def generate_html(report: ReportData) -> str:
        html_template = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{course_name} - 学生开发日志报告</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: "Microsoft YaHei", "Noto Sans SC", sans-serif; background: #f5f5f5; color: #333; line-height: 1.6; }}
        .container {{ max-width: 900px; margin: 20px auto; padding: 0 20px; }}
        .card {{ background: #fff; border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,.12); margin-bottom: 20px; overflow: hidden; }}
        .card-header {{ background: #1565c0; color: #fff; padding: 20px; text-align: center; }}
        .card-header h1 {{ font-size: 1.5rem; margin-bottom: 5px; }}
        .card-header h2 {{ font-size: 1.2rem; font-weight: normal; }}
        .card-body {{ padding: 20px; }}
        .info-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px; }}
        .info-item {{ background: #f8f9fa; padding: 12px; border-radius: 6px; }}
        .info-label {{ font-size: 0.85rem; color: #666; margin-bottom: 4px; }}
        .info-value {{ font-size: 1.1rem; font-weight: bold; color: #1565c0; }}
        .section-title {{ font-size: 1.1rem; color: #1565c0; margin-bottom: 15px; padding-bottom: 8px; border-bottom: 2px solid #e3f2fd; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 0.9rem; margin-bottom: 15px; }}
        th {{ background: #e3f2fd; padding: 10px; text-align: left; font-weight: bold; }}
        td {{ padding: 8px 10px; border-bottom: 1px solid #f0f0f0; }}
        tr:hover td {{ background: #fafafa; }}
        .badge {{ display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 0.8rem; }}
        .badge-success {{ background: #e8f5e9; color: #2e7d32; }}
        .badge-warning {{ background: #fff3e0; color: #e65100; }}
        .badge-info {{ background: #e3f2fd; color: #1565c0; }}
        .stats-row {{ display: flex; gap: 15px; flex-wrap: wrap; margin-bottom: 15px; }}
        .stat-box {{ flex: 1; min-width: 120px; background: #f8f9fa; padding: 15px; border-radius: 8px; text-align: center; }}
        .stat-value {{ font-size: 1.5rem; font-weight: bold; color: #1565c0; }}
        .stat-label {{ font-size: 0.85rem; color: #666; }}
        .no-data {{ color: #999; font-style: italic; padding: 20px; text-align: center; }}
        .footer {{ text-align: center; padding: 20px; color: #888; font-size: 0.85rem; }}
        .comment-item {{ background: #f8f9fa; padding: 12px; border-radius: 6px; margin-bottom: 10px; border-left: 3px solid #1565c0; }}
        .comment-date {{ font-size: 0.8rem; color: #888; margin-bottom: 5px; }}
        @media print {{ body {{ background: #fff; }} .card {{ box-shadow: none; }} }}
    </style>
</head>
<body>
    <div class="container">
        <div class="card">
            <div class="card-header">
                <h1>{course_name}</h1>
                <h2>学生开发日志报告</h2>
            </div>
            <div class="card-body">
                <div class="info-grid">
                    <div class="info-item">
                        <div class="info-label">姓名</div>
                        <div class="info-value">{student_name}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">学号</div>
                        <div class="info-value">{student_id}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">班级</div>
                        <div class="info-value">{class_name}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">学期</div>
                        <div class="info-value">{semester}</div>
                    </div>
                </div>
                <div class="info-item" style="text-align: center;">
                    <div class="info-label">生成时间</div>
                    <div class="info-value">{generated_at}</div>
                </div>
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">一、分支记录 ({branch_count}个分支)</h3>
                {branch_content}
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">二、提交历史 ({commit_count}次提交)</h3>
                {commit_content}
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">三、代码质量</h3>
                <div class="stats-row">
                    <div class="stat-box">
                        <div class="stat-value">{total_lines:,}</div>
                        <div class="stat-label">总代码行数</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{language_count}</div>
                        <div class="stat-label">使用语言</div>
                    </div>
                </div>
                {language_content}
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">四、Pull Request ({pr_count}个)</h3>
                {pr_content}
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">五、PR分析统计</h3>
                <div class="stats-row">
                    <div class="stat-box">
                        <div class="stat-value">{pr_total}</div>
                        <div class="stat-label">总PR数</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{pr_merged}</div>
                        <div class="stat-label">已合并</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{pr_rate}%</div>
                        <div class="stat-label">合并率</div>
                    </div>
                </div>
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">六、教师评语</h3>
                {comment_content}
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">七、在线考试成绩 ({exam_count}次考试)</h3>
                {exam_content}
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">八、考勤记录</h3>
                <div class="stats-row">
                    <div class="stat-box">
                        <div class="stat-value">{att_total}</div>
                        <div class="stat-label">总天数</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{att_rate}%</div>
                        <div class="stat-label">出勤率</div>
                    </div>
                </div>
            </div>
        </div>

        <div class="card">
            <div class="card-body">
                <h3 class="section-title">九、课程总评</h3>
                <div class="stats-row">
                    <div class="stat-box">
                        <div class="stat-value">{avg_score}</div>
                        <div class="stat-label">平均分</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{total_score}</div>
                        <div class="stat-label">总分</div>
                    </div>
                </div>
            </div>
        </div>

        <div class="footer">
            <p>报告由系统自动生成</p>
            <p>成都信息工程大学 软件工程学院</p>
        </div>
    </div>
</body>
</html>
        """

        def make_table(headers, rows, limit=10):
            if not rows:
                return '<div class="no-data">暂无数据</div>'
            table = '<table><thead><tr>'
            for h in headers:
                table += f'<th>{h}</th>'
            table += '</tr></thead><tbody>'
            for row in rows[:limit]:
                table += '<tr>'
                for cell in row:
                    table += f'<td>{cell}</td>'
                table += '</tr>'
            table += '</tbody></table>'
            if len(rows) > limit:
                table += f'<p style="color:#888;font-size:0.85rem">...还有{len(rows)-limit}条记录</p>'
            return table

        branch_content = make_table(['分支名', '保护', '最后提交时间'], [
            [b.name, '是' if b.protected else '否', b.last_commit_date[:10] if b.last_commit_date else '-']
            for b in report.branches
        ])

        commit_content = make_table(['SHA', '提交信息', '作者', '日期'], [
            [c.sha[:7], c.message[:50] + ('...' if len(c.message) > 50 else ''), c.author_name, c.date[:10] if c.date else '-']
            for c in report.commits
        ])

        language_rows = [[l['name'], f"{l['percentage']}%", f"{l['bytes']:,} bytes"] for l in report.code_quality.top_languages]
        language_content = make_table(['语言', '占比', '大小'], language_rows) if language_rows else '<div class="no-data">暂无语言数据</div>'

        pr_rows = []
        for pr in report.pull_requests[:10]:
            state = '已合并' if pr.merged_at else pr.state
            pr_rows.append([f"#{pr.number}", pr.title[:40] + ('...' if len(pr.title) > 40 else ''), state, pr.created_at[:10] if pr.created_at else '-'])
        pr_content = make_table(['#', '标题', '状态', '创建时间'], pr_rows)

        comment_items = ''
        for tc in report.teacher_comments:
            date_str = tc.created_at[:10] if tc.created_at else ''
            comment_items += f'''
            <div class="comment-item">
                <div class="comment-date">{date_str}</div>
                <div>{tc.comment}</div>
            </div>'''
        comment_content = comment_items if comment_items else '<div class="no-data">暂无评语</div>'

        exam_rows = [[e.exam_title, str(e.score), str(e.total), e.submitted_at[:10] if e.submitted_at else '-'] for e in report.exam_scores]
        exam_content = make_table(['考试名称', '得分', '总分', '提交时间'], exam_rows)

        return html_template.format(
            course_name=report.course_name,
            student_name=report.student_info.name,
            student_id=report.student_info.student_id,
            class_name=report.student_info.class_name or '未分班',
            semester=report.semester,
            generated_at=report.generated_at,
            branch_count=len(report.branches),
            branch_content=branch_content,
            commit_count=len(report.commits),
            commit_content=commit_content,
            total_lines=report.code_quality.total_lines,
            language_count=len(report.code_quality.languages),
            language_content=language_content,
            pr_count=len(report.pull_requests),
            pr_content=pr_content,
            pr_total=report.pr_analysis.total_prs,
            pr_merged=report.pr_analysis.merged_prs,
            pr_rate=report.pr_analysis.merge_rate,
            comment_content=comment_content,
            exam_count=len(report.exam_scores),
            exam_content=exam_content,
            att_total=report.attendance_summary.total_days,
            att_rate=report.attendance_summary.attendance_rate,
            avg_score=report.course_summary.avg_score,
            total_score=report.course_summary.total_score
        )

    @staticmethod
    def generate_excel(report: ReportData) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill = PatternFill("solid", fgColor="1565c0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        sub_header_fill = PatternFill("solid", fgColor="e3f2fd")
        sub_header_font = Font(bold=True, color="1565c0", size=10)

        ws_cover = wb.create_sheet("封面")
        ws_cover.column_dimensions['A'].width = 20
        ws_cover.column_dimensions['B'].width = 40

        ws_cover['A1'] = report.course_name
        ws_cover['A1'].font = Font(bold=True, size=16, color="1565c0")
        ws_cover['A2'] = "学生开发日志报告"
        ws_cover['A2'].font = Font(bold=True, size=14)
        ws_cover.merge_cells('A1:B1')
        ws_cover.merge_cells('A2:B2')

        cover_data = [
            ['姓名', report.student_info.name],
            ['学号', report.student_info.student_id],
            ['班级', report.student_info.class_name or '未分班'],
            ['学期', report.semester],
            ['生成时间', report.generated_at],
            ['', ''],
            ['GitHub数据', '已获取' if report.has_github_data else '未配置'],
            ['错误信息', report.github_error or '-']
        ]
        for i, (key, value) in enumerate(cover_data, 4):
            ws_cover.cell(row=i, column=1, value=key).font = Font(bold=True)
            ws_cover.cell(row=i, column=2, value=value)

        def create_data_sheet(name, headers, rows):
            ws = wb.create_sheet(name)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="f8f9fa")
            return ws

        branch_headers = ['分支名', '是否保护', '最后提交SHA', '最后提交时间']
        branch_rows = [[b.name, '是' if b.protected else '否', b.last_commit_sha[:7] if b.last_commit_sha else '-', b.last_commit_date[:10] if b.last_commit_date else '-'] for b in report.branches]
        create_data_sheet('分支记录', branch_headers, branch_rows)

        commit_headers = ['SHA', '提交信息', '作者', '日期', '新增行', '删除行']
        commit_rows = [[c.sha[:7], c.message, c.author_name, c.date[:10] if c.date else '-', c.additions, c.deletions] for c in report.commits]
        create_data_sheet('提交历史', commit_headers, commit_rows)

        pr_headers = ['#', '标题', '状态', '作者', '创建时间', '合并时间', '新增', '删除']
        pr_rows = [[p.number, p.title, '已合并' if p.merged_at else p.state, p.user_login, p.created_at[:10] if p.created_at else '-', p.merged_at[:10] if p.merged_at else '-', p.additions, p.deletions] for p in report.pull_requests]
        create_data_sheet('Pull Request', pr_headers, pr_rows)

        lang_headers = ['语言', '字节数', '占比']
        lang_rows = [[l['name'], l['bytes'], f"{l['percentage']}%"] for l in report.code_quality.top_languages]
        create_data_sheet('代码质量', lang_headers, lang_rows)

        comment_headers = ['日期', '评语内容']
        comment_rows = [[tc.created_at[:10] if tc.created_at else '-', tc.comment] for tc in report.teacher_comments]
        create_data_sheet('教师评语', comment_headers, comment_rows)

        exam_headers = ['考试ID', '考试名称', '得分', '总分', '提交时间']
        exam_rows = [[e.exam_id, e.exam_title, e.score, e.total, e.submitted_at[:10] if e.submitted_at else '-'] for e in report.exam_scores]
        create_data_sheet('在线考试', exam_headers, exam_rows)

        att_headers = ['日期', '状态', '备注']
        att_rows = [[a.date, a.status, a.note] for a in report.attendance_records]
        create_data_sheet('考勤记录', att_headers, att_rows)

        ws_summary = wb.create_sheet('课程总评')
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 30
        summary_data = [
            ['维度', '统计'],
            ['完成考试数', f"{report.course_summary.completed_exams}/{report.course_summary.total_exams}"],
            ['平均分', f"{report.course_summary.avg_score}分"],
            ['总分', f"{report.course_summary.total_score}分"],
            ['出勤天数', f"{report.attendance_summary.present_days}天"],
            ['出勤率', f"{report.attendance_summary.attendance_rate}%"],
            ['', ''],
            ['GitHub统计', ''],
            ['分支数', len(report.branches)],
            ['提交数', len(report.commits)],
            ['PR总数', report.pr_analysis.total_prs],
            ['合并PR数', report.pr_analysis.merged_prs],
            ['合并率', f"{report.pr_analysis.merge_rate}%"],
        ]
        for row_idx, (key, value) in enumerate(summary_data, 1):
            cell_a = ws_summary.cell(row=row_idx, column=1, value=key)
            cell_b = ws_summary.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                cell_a.font = header_font
                cell_a.fill = header_fill
                cell_b.font = header_font
                cell_b.fill = header_fill
            elif row_idx == 8:
                cell_a.font = sub_header_font
                cell_a.fill = sub_header_fill
                cell_b.fill = sub_header_fill

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def generate_zip(files: List[tuple[str, str, bytes]]) -> bytes:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename, _, content in files:
                zf.writestr(filename, content)
        return buffer.getvalue()

    @staticmethod
    def get_export_filename(report: ReportData, format: str) -> str:
        student_name = report.student_info.name
        student_id = report.student_info.student_id
        date_str = datetime.now().strftime("%Y%m%d")

        if format == "pdf":
            return f"开发日志报告_{student_name}_{student_id}_{date_str}.pdf"
        elif format == "html":
            return f"开发日志报告_{student_name}_{student_id}_{date_str}.html"
        elif format == "excel":
            return f"开发日志报告_{student_name}_{student_id}_{date_str}.xlsx"
        else:
            return f"开发日志报告_{student_name}_{student_id}_{date_str}.{format}"
