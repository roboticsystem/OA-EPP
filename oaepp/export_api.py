"""
学生开发日志导出 API（F-T-010）。
9 维度数据从平台 MySQL 拉取 + PDF/HTML/Excel 导出 + 批量班级导出 + 审计日志。
"""
import io
import json
import os
import zipfile
from datetime import datetime
from urllib.parse import quote

from database import db

# ── 学生信息查询 ────────────────────────────────────────────────────────────

_STUDENT_QUERY = """
    SELECT u.id AS user_id, u.student_no, u.full_name, u.email,
           s.class_name, gb.github_username, gb.github_name
    FROM users u
    JOIN students s ON s.user_id = u.id
    LEFT JOIN github_bindings gb ON gb.student_user_id = u.id AND gb.verify_status = 'approved'
    WHERE u.role = 'student' AND u.is_active = 1
"""


def _get_student(user_id: int) -> dict:
    with db() as conn:
        cur = conn.cursor()
        cur.execute(_STUDENT_QUERY + " AND u.id = %s", (user_id,))
        return cur.fetchone() or {}


def _get_student_by_no(student_no: str) -> dict:
    with db() as conn:
        cur = conn.cursor()
        cur.execute(_STUDENT_QUERY + " AND u.student_no = %s", (student_no,))
        return cur.fetchone() or {}


def _get_user_id_by_no(student_no: str) -> int:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM users WHERE student_no=%s AND role='student'", (student_no,))
        r = cur.fetchone()
        return r["id"] if r else None


# ── 9 维度数据收集 ──────────────────────────────────────────────────────────


def _collect_branches(user_id: int, course_id: int = None) -> list:
    """分支记录 — 来自 submissions（每次提交视作一个分支点的产出）。"""
    with db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT sb.id, a.title AS assignment_title, c.name AS course_name,
                   sb.version_no, sb.grading_status, sb.is_late, sb.submitted_at
            FROM submissions sb
            JOIN assignments a ON a.id = sb.assignment_id
            JOIN chapters ch ON ch.id = a.chapter_id
            JOIN courses c ON c.id = ch.course_id
            WHERE sb.student_user_id = %s
        """
        params = [user_id]
        if course_id:
            sql += " AND c.id = %s"
            params.append(course_id)
        sql += " ORDER BY sb.submitted_at"
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


def _collect_commits(user_id: int, course_id: int = None) -> list:
    """提交历史 — 来自 submissions 的版本记录，模拟 commit log。"""
    branches = _collect_branches(user_id, course_id)
    commits = []
    for b in branches:
        commits.append({
            "repo_name": b.get("course_name", ""),
            "branch_name": b.get("assignment_title", ""),
            "version_no": b.get("version_no", 1),
            "message": f"提交第{b.get('version_no',1)}版 — {b.get('assignment_title','')}",
            "committed_at": str(b.get("submitted_at", "")),
            "additions": 0,
            "deletions": 0,
            "is_late": b.get("is_late", 0),
        })
    return commits


def _collect_code_quality(user_id: int, course_id: int = None) -> list:
    """代码质量 — 来自 pr_records 的 quality_score + commitlint 配置评估。"""
    with db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT pr.id, pr.issue_no, pr.pr_number, pr.quality_score, pr.pr_state,
                   c.name AS course_name
            FROM pr_records pr
            JOIN courses c ON c.id = pr.course_id
            WHERE pr.student_user_id = %s
        """
        params = [user_id]
        if course_id:
            sql += " AND pr.course_id = %s"
            params.append(course_id)
        sql += " ORDER BY pr.merged_at DESC, pr.id DESC"
        cur.execute(sql, params)
        rows = cur.fetchall()
    metrics = []
    for r in rows:
        metrics.append({
            "repo_name": r.get("course_name", ""),
            "metric_name": "PR质量评分",
            "metric_value": float(r.get("quality_score") or 0),
            "recorded_at": "",
            "detail": f"Issue#{r.get('issue_no')} PR#{r.get('pr_number')} [{r.get('pr_state')}]",
        })
    return metrics


def _collect_prs(user_id: int, course_id: int = None) -> list:
    """PR 情况 — 来自 pr_records。"""
    with db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT pr.id, c.name AS repo_name, pr.issue_no, pr.pr_number,
                   pr.pr_state AS state, pr.quality_score, pr.merged_at,
                   NULL AS created_at
            FROM pr_records pr
            JOIN courses c ON c.id = pr.course_id
            WHERE pr.student_user_id = %s
        """
        params = [user_id]
        if course_id:
            sql += " AND pr.course_id = %s"
            params.append(course_id)
        sql += " ORDER BY pr.merged_at DESC"
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


def _collect_pr_analysis(user_id: int, course_id: int = None) -> list:
    """PR 分析 — 对 PR 做统计摘要。"""
    prs = _collect_prs(user_id, course_id)
    analysis = []
    for p in prs:
        analysis.append({
            "pr_number": p.get("pr_number"),
            "issue_no": p.get("issue_no"),
            "title": f"Issue#{p.get('issue_no')} → PR#{p.get('pr_number')}",
            "state": p.get("state"),
            "quality_score": float(p.get("quality_score") or 0),
            "merged": p.get("state") == "merged",
            "merged_at": str(p.get("merged_at") or ""),
        })
    return analysis


def _collect_teacher_comments(user_id: int, course_id: int = None) -> list:
    """教师评语 — 来自 feedbacks 表。"""
    with db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT f.source_type, f.source_id, f.content, f.created_at,
                   u.full_name AS teacher_name
            FROM feedbacks f
            LEFT JOIN users u ON u.id = f.teacher_user_id
            WHERE f.student_user_id = %s
        """
        params = [user_id]
        if course_id:
            sql += " AND f.source_id IN (SELECT id FROM exams WHERE course_id=%s)" \
                   " OR f.source_type IN ('assignment','pr','manual')"
        sql += " ORDER BY f.created_at DESC"
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


def _collect_exams(user_id: int, course_id: int = None) -> list:
    """在线考试 — 来自 exams + exam_attempts。"""
    with db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT e.title, e.exam_type, ea.total_score AS score, ea.status,
                   ea.submitted_at, c.name AS course_name
            FROM exam_attempts ea
            JOIN exams e ON e.id = ea.exam_id
            JOIN courses c ON c.id = e.course_id
            WHERE ea.student_user_id = %s
        """
        params = [user_id]
        if course_id:
            sql += " AND e.course_id = %s"
            params.append(course_id)
        sql += " ORDER BY ea.submitted_at"
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


def _collect_attendance(user_id: int, course_id: int = None) -> dict:
    """考勤 — 来自 attendance_records。"""
    with db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT ar.status, ar.checkin_at, c.name AS course_name
            FROM attendance_records ar
            JOIN courses c ON c.id = ar.course_id
            WHERE ar.student_user_id = %s
        """
        params = [user_id]
        if course_id:
            sql += " AND ar.course_id = %s"
            params.append(course_id)
        sql += " ORDER BY ar.checkin_at"
        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]
    total = len(rows)
    present = sum(1 for r in rows if r["status"] == "present")
    return {
        "records": rows,
        "total": total,
        "present": present,
        "rate": f"{(present / total * 100):.1f}%" if total > 0 else "N/A",
    }


def _collect_course_scores(user_id: int, course_id: int = None) -> list:
    """课程得分 — 来自 score_items（按 attendance/exam/code/pr 分类）。"""
    with db() as conn:
        cur = conn.cursor()
        sql = """
            SELECT si.score_type AS category, si.score, si.scored_at,
                   c.name AS course_name, c.term AS semester,
                   gwc.attendance_weight, gwc.exam_weight, gwc.code_weight, gwc.pr_weight
            FROM score_items si
            JOIN courses c ON c.id = si.course_id
            LEFT JOIN grade_weight_configs gwc ON gwc.course_id = si.course_id
            WHERE si.student_user_id = %s
        """
        params = [user_id]
        if course_id:
            sql += " AND si.course_id = %s"
            params.append(course_id)
        sql += " ORDER BY si.scored_at"
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


# ── 综合数据收集 ────────────────────────────────────────────────────────────


def collect_all(student_no: str, course_id: int = None) -> dict:
    """收集指定学生的全部 9 维度数据。"""
    student = _get_student_by_no(student_no)
    if not student:
        raise ValueError(f"学生 {student_no} 不存在")
    uid = student["user_id"]
    return {
        "student": student,
        "branches": _collect_branches(uid, course_id),
        "commits": _collect_commits(uid, course_id),
        "code_quality": _collect_code_quality(uid, course_id),
        "prs": _collect_prs(uid, course_id),
        "pr_analysis": _collect_pr_analysis(uid, course_id),
        "teacher_comments": _collect_teacher_comments(uid, course_id),
        "exams": _collect_exams(uid, course_id),
        "attendance": _collect_attendance(uid, course_id),
        "course_scores": _collect_course_scores(uid, course_id),
    }


# ── 审计日志 ─────────────────────────────────────────────────────────────────


def _audit(actor_student_no: str, target_student_no: str, fmt: str, class_name: str = ""):
    """写入操作审计日志到 audit_logs 表。"""
    try:
        actor_id = _get_user_id_by_no(actor_student_no)
        target_id = _get_user_id_by_no(target_student_no)
        with db() as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO audit_logs (actor_user_id, action, target_type, target_id, detail_json, action_at) "
                "VALUES (%s,%s,%s,%s,%s,NOW())",
                (actor_id or 0, "export_report", "student", target_id or 0,
                 json.dumps({"format": fmt, "class_name": class_name, "exporter_no": actor_student_no})))
    except Exception:
        pass


def get_audit_logs(limit: int = 100) -> list:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT al.id, u.student_no AS exporter_no, al.action, al.target_type,
                   al.target_id, al.detail_json, al.action_at
            FROM audit_logs al
            LEFT JOIN users u ON u.id = al.actor_user_id
            WHERE al.action = 'export_report'
            ORDER BY al.action_at DESC LIMIT %s
        """, (limit,))
        return [dict(r) for r in cur.fetchall()]


# ── 学生搜索和班级查询 ──────────────────────────────────────────────────────


def search_students(q: str, limit: int = 20) -> list:
    with db() as conn:
        cur = conn.cursor()
        like = f"%{q}%"
        cur.execute(_STUDENT_QUERY + """
            AND (u.student_no LIKE %s OR u.full_name LIKE %s OR u.email LIKE %s)
            ORDER BY u.student_no LIMIT %s
        """, (like, like, like, limit))
        return [dict(r) for r in cur.fetchall()]


def get_all_classes() -> list:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT class_name FROM students WHERE class_name != '' ORDER BY class_name")
        return [r["class_name"] for r in cur.fetchall()]


def get_students_by_class(class_name: str) -> list:
    with db() as conn:
        cur = conn.cursor()
        cur.execute(_STUDENT_QUERY + " AND s.class_name = %s ORDER BY u.student_no", (class_name,))
        return [dict(r) for r in cur.fetchall()]


# ── HTML 导出 ────────────────────────────────────────────────────────────────


def _render_html(data: dict) -> str:
    s = data["student"]
    header = f"""
    <div class="cover">
      <h1>学生开发日志报告</h1>
      <p>学号：{s.get('student_no','')} | 姓名：{s.get('full_name','')} | 班级：{s.get('class_name','')}</p>
      <p>GitHub：{s.get('github_username','未绑定')} ({s.get('github_name','')})</p>
      <p>邮箱：{s.get('email','')}</p>
      <p>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
    """

    def tbl(title, rows, columns, key_map=None):
        if not rows:
            return f"<h2>{title}</h2><p>暂无数据</p>"
        html = f"<h2>{title}</h2><table border='1' cellpadding='6' cellspacing='0' style='border-collapse:collapse;width:100%'>"
        html += "<tr>" + "".join(f"<th>{c}</th>" for c in columns) + "</tr>"
        for r in rows:
            html += "<tr>" + "".join(f"<td>{r.get(c,'') if key_map is None else r.get(key_map.get(c,c),'')}</td>" for c in columns) + "</tr>"
        html += "</table>"
        return html

    sections = [
        tbl("1. 分支记录 / 作业提交", data["branches"],
            ["assignment_title", "course_name", "version_no", "grading_status", "is_late", "submitted_at"]),
        tbl("2. 提交历史", data["commits"],
            ["repo_name", "branch_name", "version_no", "message", "committed_at", "is_late"]),
        tbl("3. 代码质量 (PR 评分)", data["code_quality"],
            ["repo_name", "metric_name", "metric_value", "detail"]),
        tbl("4. PR 情况", data["prs"],
            ["repo_name", "issue_no", "pr_number", "state", "quality_score", "merged_at"]),
        tbl("5. PR 分析", data["pr_analysis"],
            ["pr_number", "issue_no", "title", "state", "quality_score", "merged", "merged_at"]),
        tbl("6. 教师评语", data["teacher_comments"],
            ["teacher_name", "source_type", "content", "created_at"]),
        tbl("7. 在线考试", data["exams"],
            ["course_name", "title", "exam_type", "score", "status", "submitted_at"]),
        tbl("9. 课程得分", data["course_scores"],
            ["course_name", "semester", "category", "score"]),
    ]

    att = data["attendance"]
    attendance_html = f"<h2>8. 考勤记录</h2><p>总课时：{att.get('total',0)} | 出勤：{att.get('present',0)} | 出勤率：{att.get('rate','N/A')}</p>"
    if att.get("records"):
        attendance_html += tbl("", att["records"], ["course_name", "status", "checkin_at"])

    sections.insert(7, attendance_html)

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>开发日志报告 - {s.get('full_name','')}</title>
<style>
  body {{ font-family: 'Microsoft YaHei',sans-serif; margin:40px; color:#333; }}
  h1 {{ color:#1a56db; border-bottom:3px solid #1a56db; padding-bottom:10px; }}
  h2 {{ color:#2563eb; margin-top:30px; }}
  table {{ margin-bottom:20px; }}
  th {{ background:#2563eb; color:#fff; }}
  tr:nth-child(even) {{ background:#f0f4ff; }}
  .cover {{ text-align:center; margin-bottom:40px; }}
</style></head>
<body>
  {header}
  {"".join(sections)}
</body></html>"""


# ── PDF 导出 ─────────────────────────────────────────────────────────────────


def _render_pdf(data: dict) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError("请安装 fpdf 库: pip install fpdf2")

    s = data["student"]
    pdf = FPDF()
    pdf.add_page()

    font_paths = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
    ]
    font_ok = False
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdf.add_font("CJK", "", fp, uni=True)
                pdf.add_font("CJK", "B", fp, uni=True)
                font_ok = True
                break
            except Exception:
                continue
    if not font_ok:
        pdf.add_font("CJK", "", "Helvetica")
        pdf.add_font("CJK", "B", "Helvetica")

    pdf.set_font("CJK", "B", 16)
    pdf.cell(0, 10, "学生开发日志报告", ln=True, align="C")
    pdf.set_font("CJK", "", 10)
    pdf.cell(0, 7, f"学号: {s.get('student_no','')}  姓名: {s.get('full_name','')}  班级: {s.get('class_name','')}", ln=True, align="C")
    pdf.cell(0, 7, f"GitHub: {s.get('github_username','未绑定')}  生成: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(5)

    sections_def = [
        ("1. 分支记录", data["branches"], ["assignment_title", "course_name", "version_no", "grading_status"]),
        ("2. 提交历史", data["commits"], ["repo_name", "branch_name", "version_no", "message", "committed_at"]),
        ("3. 代码质量", data["code_quality"], ["repo_name", "metric_name", "metric_value", "detail"]),
        ("4. PR 情况", data["prs"], ["repo_name", "issue_no", "pr_number", "state", "quality_score"]),
        ("5. PR 分析", data["pr_analysis"], ["title", "state", "quality_score", "merged"]),
        ("6. 教师评语", data["teacher_comments"], ["teacher_name", "source_type", "content", "created_at"]),
        ("7. 在线考试", data["exams"], ["course_name", "title", "exam_type", "score", "status"]),
        ("8. 考勤", [], []),
        ("9. 课程得分", data["course_scores"], ["course_name", "semester", "category", "score"]),
    ]

    for title, rows, cols in sections_def:
        pdf.set_font("CJK", "B", 11)
        if title == "8. 考勤":
            att = data["attendance"]
            pdf.cell(0, 7, f"8. 考勤记录  (总: {att.get('total',0)}  出勤: {att.get('present',0)}  出勤率: {att.get('rate','N/A')})", ln=True)
            rows = att.get("records", [])
            cols = ["course_name", "status", "checkin_at"]
        else:
            pdf.cell(0, 7, title, ln=True)

        pdf.set_font("CJK", "", 8)
        if rows:
            col_w = 190 // max(len(cols), 1)
            for c in cols:
                pdf.cell(col_w, 6, c[:20], border=1)
            pdf.ln()
            for r in rows:
                for c in cols:
                    val = str(r.get(c, ""))[:25]
                    pdf.cell(col_w, 6, val, border=1)
                pdf.ln()
        else:
            pdf.cell(0, 6, "  暂无数据", ln=True)
        pdf.ln(3)

    return pdf.output()


# ── Excel 导出 ───────────────────────────────────────────────────────────────


def _render_excel(data: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报告封面"
    s = data["student"]
    ws.merge_cells("A1:D1")
    ws["A1"] = "学生开发日志报告"
    ws["A1"].font = Font(bold=True, size=16, color="1a56db")
    ws["A1"].alignment = Alignment(horizontal="center")
    info_lines = [
        ("学号", s.get("student_no", "")), ("姓名", s.get("full_name", "")),
        ("班级", s.get("class_name", "")), ("GitHub", s.get("github_username", "未绑定")),
        ("邮箱", s.get("email", "")),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(info_lines, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    hdr_fill = PatternFill("solid", fgColor="2563eb")
    hdr_font = Font(bold=True, color="FFFFFF")

    def _write_sheet(title, rows, columns):
        sheet = wb.create_sheet(title=title[:31])
        for ci, c in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=ci, value=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for ri, r in enumerate(rows or [], 2):
            for ci, c in enumerate(columns, 1):
                sheet.cell(row=ri, column=ci, value=str(r.get(c, "")))

    _write_sheet("分支记录", data["branches"], ["assignment_title", "course_name", "version_no", "grading_status", "is_late", "submitted_at"])
    _write_sheet("提交历史", data["commits"], ["repo_name", "branch_name", "version_no", "message", "committed_at", "is_late"])
    _write_sheet("代码质量", data["code_quality"], ["repo_name", "metric_name", "metric_value", "detail"])
    _write_sheet("PR情况", data["prs"], ["repo_name", "issue_no", "pr_number", "state", "quality_score", "merged_at"])
    _write_sheet("PR分析", data["pr_analysis"], ["pr_number", "issue_no", "title", "state", "quality_score", "merged", "merged_at"])
    _write_sheet("教师评语", data["teacher_comments"], ["teacher_name", "source_type", "content", "created_at"])
    _write_sheet("在线考试", data["exams"], ["course_name", "title", "exam_type", "score", "status", "submitted_at"])
    _write_sheet("课程得分", data["course_scores"], ["course_name", "semester", "category", "score"])

    ws_att = wb.create_sheet(title="考勤记录")
    att = data["attendance"]
    ws_att.cell(row=1, column=1, value=f"总课时: {att.get('total',0)}  出勤: {att.get('present',0)}  出勤率: {att.get('rate','N/A')}")
    for ci, c in enumerate(["course_name", "status", "checkin_at"], 1):
        cell = ws_att.cell(row=2, column=ci, value=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for ri, r in enumerate(att.get("records", []), 3):
        ws_att.cell(row=ri, column=1, value=str(r.get("course_name", "")))
        ws_att.cell(row=ri, column=2, value=r.get("status", ""))
        ws_att.cell(row=ri, column=3, value=str(r.get("checkin_at", "")))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 导出入口 ─────────────────────────────────────────────────────────────────


def export_report(student_no: str, fmt: str = "html", exporter_no: str = "") -> bytes:
    """导出单个学生报告。"""
    data = collect_all(student_no)
    _audit(actor_student_no=exporter_no, target_student_no=student_no, fmt=fmt)
    if fmt == "html":
        return _render_html(data).encode("utf-8")
    elif fmt == "pdf":
        return _render_pdf(data)
    elif fmt == "excel":
        return _render_excel(data)
    else:
        raise ValueError(f"不支持的导出格式: {fmt}")


def export_batch_by_class(class_name: str, fmt: str = "excel", exporter_no: str = "") -> bytes:
    """按班级批量导出 ZIP。"""
    students = get_students_by_class(class_name)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in students:
            try:
                data = collect_all(s["student_no"])
                _audit(actor_student_no=exporter_no, target_student_no=s["student_no"], fmt=fmt, class_name=class_name)
                if fmt == "html":
                    content = _render_html(data).encode("utf-8")
                    ext = "html"
                elif fmt == "pdf":
                    content = _render_pdf(data)
                    ext = "pdf"
                else:
                    content = _render_excel(data)
                    ext = "xlsx"
                zf.writestr(f"{s['student_no']}_{s['full_name']}.{ext}", content)
            except Exception as e:
                zf.writestr(f"{s['student_no']}_{s['full_name']}_error.txt", str(e).encode("utf-8"))
    buf.seek(0)
    return buf.read()
