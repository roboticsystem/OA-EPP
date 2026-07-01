"""F-S-032 成绩导出（学生端）— GradeExportState

对应原型：prototype/grades.html（"下载成绩单 Excel" 按钮）
对应页面：oaepp/pages/grades.py
对应需求：docs/第10章_软件产品介绍.md F-S-032

验收要点（来自需求文档）：
- 学生一键下载本人全期成绩单（.xlsx 格式）
- 内容必含：学号、姓名、各任务得分、总评成绩、提交时间、批改时间
- 仅包含本人数据
- 文件名格式：学号_姓名_成绩单_课程名称.xlsx

协作规范（来自 学生功能开发指南.md）：
- 独立 rx.State，不继承 GlobalState，不修改全局状态
- 通过 oaepp.database.db_sync() 使用公共数据库连接池，不自建连接
- 通过 AuthState 只读获取当前登录用户
"""

import datetime
import io
from typing import Any, Dict, List, Optional

try:
    import reflex as rx
except Exception:
    rx = None

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


# ── 数据查询与 Excel 构建（纯函数，便于单元测试） ─────────────────────────

def _compute_total_score(course_data: dict) -> Optional[float]:
    """按 grade_weight_configs 权重计算课程总评成绩。

    无权重配置时退化为简单求和；无任何 score_items 时返回 None。
    """
    scores_by_type = {"attendance": 0.0, "exam": 0.0, "code": 0.0, "pr": 0.0}
    counts_by_type = {"attendance": 0, "exam": 0, "code": 0, "pr": 0}

    for si in course_data.get("score_items", []):
        st = si["score_type"]
        if st in scores_by_type:
            scores_by_type[st] += si["score"]
            counts_by_type[st] += 1

    weights = course_data.get("weights")
    if not weights:
        items = course_data.get("score_items", [])
        return sum(si["score"] for si in items) if items else None

    weighted_total = 0.0
    for dim in ("attendance", "exam", "code", "pr"):
        avg = scores_by_type[dim] / counts_by_type[dim] if counts_by_type[dim] > 0 else 0.0
        weight_pct = weights.get(dim, 0.0) / 100.0
        weighted_total += avg * weight_pct
    return round(weighted_total, 2)


def _query_student_grades(student_no: str) -> Optional[Dict[str, Any]]:
    """查询学生的全部成绩数据；返回 None 表示学生不存在。

    使用公共数据库连接池 oaepp.database.db_sync()。
    """
    try:
        from database import db_sync
    except ImportError:
        from oaepp.database import db_sync

    with db_sync() as cur:
        cur.execute(
            """SELECT id, student_no, full_name, email
               FROM users
               WHERE student_no = %s AND role = 'student'""",
            (student_no,),
        )
        user = cur.fetchone()
        if not user:
            return None

        student_info = {
            "user_id": user["id"],
            "student_no": user["student_no"],
            "full_name": user["full_name"],
            "email": user["email"],
        }

        cur.execute(
            """SELECT c.id, c.code, c.name, c.term
               FROM courses c
               JOIN enrollments e ON e.course_id = c.id
               WHERE e.student_user_id = %s
               ORDER BY c.term, c.code""",
            (user["id"],),
        )
        course_rows = cur.fetchall()

        courses: List[dict] = []
        for c_row in course_rows:
            course_data = {
                "course_id": c_row["id"],
                "course_code": c_row["code"],
                "course_name": c_row["name"],
                "course_term": c_row["term"],
                "tasks": [],
                "score_items": [],
                "total_score": None,
            }

            cur.execute(
                """SELECT
                    a.id AS assignment_id,
                    a.title AS assignment_title,
                    a.deadline,
                    s.id AS submission_id,
                    s.version_no,
                    s.submitted_at,
                    s.is_late,
                    s.grading_status,
                    gr.id AS grading_id,
                    gr.attendance_score,
                    gr.exam_score,
                    gr.code_score,
                    gr.pr_score,
                    gr.total_score AS grading_total,
                    gr.comment_md,
                    gr.graded_at
                   FROM assignments a
                   LEFT JOIN submissions s
                     ON s.assignment_id = a.id
                     AND s.student_user_id = %s
                   LEFT JOIN grading_records gr
                     ON gr.submission_id = s.id
                   WHERE a.course_id = %s
                   ORDER BY a.deadline ASC""",
                (user["id"], c_row["id"]),
            )
            for t_row in cur.fetchall():
                course_data["tasks"].append({
                    "assignment_id": t_row["assignment_id"],
                    "assignment_title": t_row["assignment_title"],
                    "deadline": t_row["deadline"],
                    "submission_id": t_row["submission_id"],
                    "version_no": t_row["version_no"],
                    "submitted_at": t_row["submitted_at"],
                    "is_late": bool(t_row["is_late"]) if t_row["is_late"] is not None else False,
                    "grading_status": t_row["grading_status"],
                    "grading_id": t_row["grading_id"],
                    "attendance_score": t_row["attendance_score"],
                    "exam_score": t_row["exam_score"],
                    "code_score": t_row["code_score"],
                    "pr_score": t_row["pr_score"],
                    "grading_total": t_row["grading_total"],
                    "comment_md": t_row["comment_md"],
                    "graded_at": t_row["graded_at"],
                })

            cur.execute(
                """SELECT si.score_type, si.score, si.scored_at
                   FROM score_items si
                   WHERE si.course_id = %s AND si.student_user_id = %s
                   ORDER BY si.scored_at ASC""",
                (c_row["id"], user["id"]),
            )
            for si_row in cur.fetchall():
                course_data["score_items"].append({
                    "score_type": si_row["score_type"],
                    "score": float(si_row["score"]),
                    "scored_at": si_row["scored_at"],
                })

            cur.execute(
                """SELECT attendance_weight, exam_weight, code_weight, pr_weight
                   FROM grade_weight_configs
                   WHERE course_id = %s
                   LIMIT 1""",
                (c_row["id"],),
            )
            gw = cur.fetchone()
            if gw:
                course_data["weights"] = {
                    "attendance": float(gw["attendance_weight"]),
                    "exam": float(gw["exam_weight"]),
                    "code": float(gw["code_weight"]),
                    "pr": float(gw["pr_weight"]),
                }
            else:
                course_data["weights"] = None

            course_data["total_score"] = _compute_total_score(course_data)
            courses.append(course_data)

    return {"student_info": student_info, "courses": courses}


def _fmt_dt(value, default: str = "") -> str:
    """统一格式化 datetime → 字符串，None/空 → default。"""
    if value is None or value == "":
        return default
    if isinstance(value, str):
        return value
    try:
        return value.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(value)


def _build_workbook(data: dict) -> bytes:
    """生成 Excel 工作簿（多 sheet，每门课一个）并返回 bytes。

    列顺序对齐验收口径：序号 / 学号 / 姓名 / 任务名称 / 任务类型 / 得分 /
    提交时间 / 批改时间 / 迟交 / 评语。学号、姓名作为独立列，便于二次处理。
    """
    if openpyxl is None:
        raise ImportError("openpyxl 未安装，无法生成 Excel 文件")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    student_info = data["student_info"]
    courses = data["courses"]
    student_no = student_info["student_no"]
    full_name = student_info["full_name"]

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    title_font = Font(name="微软雅黑", size=14, bold=True)

    headers = [
        "序号", "学号", "姓名", "任务名称", "任务类型",
        "得分", "提交时间", "批改时间", "迟交", "评语",
    ]
    col_widths = [6, 14, 12, 28, 10, 8, 18, 18, 6, 28]
    last_col_letter = get_column_letter(len(headers))

    for course in courses:
        course_name = course["course_name"]
        sheet_name = course_name[:31] or f"课程{course['course_id']}"
        ws = wb.create_sheet(title=sheet_name)

        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"].value = f"{course_name} — 成绩单"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"].value = (
            f"学号：{student_no}    姓名：{full_name}"
            f"    学期：{course['course_term']}    课程编号：{course['course_code']}"
        )
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A2"].font = Font(name="微软雅黑", size=10)

        header_row = 4
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row = header_row + 1
        seq = 1

        for task in course["tasks"]:
            score_value = task.get("grading_total")
            if score_value is not None:
                score_value = float(score_value)

            submitted_at = _fmt_dt(task.get("submitted_at"), default="未提交")
            graded_at = _fmt_dt(task.get("graded_at"), default="未批改")
            is_late = "是" if task.get("is_late") else "否"
            comment = task.get("comment_md") or ""

            values = [
                seq, student_no, full_name, task["assignment_title"], "作业",
                score_value, submitted_at, graded_at, is_late, comment,
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = cell_alignment if col_idx != len(headers) else cell_alignment_left
                cell.font = Font(name="微软雅黑", size=10)
            row += 1
            seq += 1

        type_label_map = {"attendance": "考勤", "exam": "考试", "code": "代码", "pr": "PR"}
        for si in course["score_items"]:
            task_type = type_label_map.get(si["score_type"], si["score_type"])
            scored_at = _fmt_dt(si.get("scored_at"))
            values = [
                seq, student_no, full_name, f"{task_type}成绩", task_type,
                si["score"], "", scored_at, "", "",
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = cell_alignment
                cell.font = Font(name="微软雅黑", size=10)
            row += 1
            seq += 1

        # 总评行
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        total_label_cell = ws.cell(row=row, column=1, value="总评成绩（按权重计算）")
        total_label_cell.font = Font(name="微软雅黑", size=11, bold=True)
        total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_label_cell.border = thin_border
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = thin_border

        total_score = course.get("total_score")
        total_cell = ws.cell(row=row, column=6, value=total_score if total_score is not None else "N/A")
        total_cell.font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
        total_cell.alignment = cell_alignment
        total_cell.border = thin_border
        for c in range(7, len(headers) + 1):
            ws.cell(row=row, column=c).border = thin_border

        # 权重说明
        row += 1
        weights = course.get("weights")
        if weights:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            weight_text = (
                f"权重配置：出勤 {weights['attendance']}%  |  "
                f"考试 {weights['exam']}%  |  "
                f"代码 {weights['code']}%  |  "
                f"PR {weights['pr']}%"
            )
            ws.cell(row=row, column=1, value=weight_text).font = Font(
                name="微软雅黑", size=9, color="666666"
            )

        # 导出时间
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        export_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row, column=1, value=f"导出时间：{export_time}").font = Font(
            name="微软雅黑", size=9, color="999999"
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_filename(student_info: dict, courses: list, fallback_student_no: str) -> str:
    """按规范生成文件名：学号_姓名_成绩单_课程名称.xlsx。

    - 单课程：学号_姓名_成绩单_课程名称.xlsx
    - 多课程：学号_姓名_成绩单_全部课程.xlsx（兜底，避免误导只显示第一门）
    - 无课程：学号_姓名_成绩单.xlsx
    """
    sno = (student_info or {}).get("student_no") or fallback_student_no
    name = (student_info or {}).get("full_name") or ""
    if not courses:
        return f"{sno}_{name}_成绩单.xlsx" if name else f"{sno}_成绩单.xlsx"
    if len(courses) == 1:
        course_name = courses[0]["course_name"]
        return (
            f"{sno}_{name}_成绩单_{course_name}.xlsx"
            if name else f"{sno}_成绩单_{course_name}.xlsx"
        )
    return f"{sno}_{name}_成绩单_全部课程.xlsx" if name else f"{sno}_成绩单_全部课程.xlsx"


# ── Reflex State（仅在 reflex 可用时定义；独立 State，不继承全局 State） ──

GradeExportState = None
if rx is not None:
    class GradeExportState(rx.State):
        """学生成绩页面 State（F-S-032 成绩导出）

        独立 State，不修改全局 GlobalState；通过 AuthState 只读获取当前用户。
        """

        # ── 隔离字段（TDD TC04 要求） ────────────────────────────────
        current_user_id: Optional[int] = None
        current_student_no: str = ""

        # ── UI 反馈 ──────────────────────────────────────────────────
        is_exporting: bool = False
        export_error: str = ""
        export_filename: str = ""

        def get_export_filename(self, student_no: str = "") -> str:
            """返回上次导出的文件名；若尚未导出且传入了学号，回退为占位名。

            兼容 TDD TC03（get_export_filename 方法签名）与
            app.py 早期 Starlette 路由调用方式。
            """
            if self.export_filename:
                return self.export_filename
            sno = student_no or self.current_student_no
            return f"{sno}_成绩单.xlsx" if sno else "成绩单.xlsx"

        async def export_my_grades(self):
            """导出当前登录学生的成绩单 .xlsx 并触发浏览器下载。"""
            self.is_exporting = True
            self.export_error = ""

            student_no = ""
            user_id: Optional[int] = None
            try:
                try:
                    from states.auth import AuthState
                except ImportError:
                    from oaepp.states.auth import AuthState
                auth = await self.get_state(AuthState)
                student_no = getattr(auth, "current_student_no", "") or ""
                user_id = getattr(auth, "current_user_id", None)
            except Exception:
                student_no = ""
                user_id = None

            self.current_student_no = student_no
            self.current_user_id = user_id

            if not student_no:
                self.export_error = "未检测到登录用户，请先登录"
                self.is_exporting = False
                return

            try:
                data = _query_student_grades(student_no)
                if data is None:
                    self.export_error = f"未找到学生 {student_no} 的成绩数据"
                    self.is_exporting = False
                    return
                file_bytes = _build_workbook(data)
                filename = _build_filename(data["student_info"], data["courses"], student_no)
                self.export_filename = filename
                self.is_exporting = False
                return rx.download(data=file_bytes, filename=filename)
            except Exception as e:
                self.export_error = f"导出失败: {e}"
                self.is_exporting = False
                return
