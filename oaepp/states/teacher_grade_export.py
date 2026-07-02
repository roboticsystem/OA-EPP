"""F-T-008 教师成绩导出 — TeacherGradeExportState

Reflex State，提供教师成绩导出的状态管理：
- class_filter / course_filter: 筛选条件
- is_exporting: 导出状态标志
- EXPORT_COLUMNS: 导出列定义（学号/姓名/GitHub用户名/各项成绩/总分）
- get_export_filename(): 文件名生成
- compute_total_score(): 总分自动计算
- audit_log: 审计日志记录
"""

import logging
from typing import Any, List, Optional

try:
    import reflex as rx
except Exception:
    rx = None

logger = logging.getLogger("teacher_grade_export")


def _compute_total_score(row: dict, weights: Optional[dict] = None) -> Optional[float]:
    """根据权重公式自动计算总评成绩"""
    if not weights:
        weights = {"attendance": 0.15, "exam": 0.30, "code": 0.35, "pr": 0.20}
    total = 0.0
    has_any = False
    for k, w in weights.items():
        val = row.get(k)
        if val is None:
            continue
        try:
            v = float(val)
            has_any = True
            total += v * float(w)
        except (TypeError, ValueError):
            continue
    return round(total, 2) if has_any else None


def _compute_grade(score: float) -> str:
    """根据分数计算等级"""
    if score is None:
        return ""
    try:
        s = float(score)
    except (TypeError, ValueError):
        return ""
    if s >= 90:
        return "A"
    if s >= 80:
        return "B"
    if s >= 70:
        return "C"
    if s >= 60:
        return "D"
    return "F"


def _load_filter_options() -> dict:
    """从数据库加载筛选选项（班级/课程/学期）"""
    try:
        from oaepp.database import db_sync
    except ImportError:
        try:
            from database import db_sync
        except ImportError:
            logger.warning("db_sync not available, returning empty filters")
            return {"classes": [], "courses": [], "terms": []}
    try:
        with db_sync() as cur:
            cur.execute(
                "SELECT DISTINCT s.class_name FROM students s "
                "JOIN users u ON u.id = s.user_id "
                "WHERE u.role='student' AND s.class_name!='' ORDER BY s.class_name"
            )
            classes = [r["class_name"] for r in cur.fetchall()]

            cur.execute("SELECT id, name, term FROM courses ORDER BY id")
            courses = [{"id": r["id"], "title": r["name"]} for r in cur.fetchall()]

            cur.execute("SELECT DISTINCT term FROM courses WHERE term!='' ORDER BY term")
            terms = [r["term"] for r in cur.fetchall()]
        return {"classes": classes, "courses": courses, "terms": terms}
    except Exception as e:
        logger.error("Failed to load filter options: %s", e)
        return {"classes": [], "courses": [], "terms": []}


TeacherGradeExportState = None
if rx is not None:
    class TeacherGradeExportState(rx.State):
        """教师成绩导出状态管理

        对齐 prototype/admin_grades.html 原型 成绩单导出 Tab：
        - 筛选导出范围（班级/课程/学期）
        - 数据预览与手动修正个别单元格
        - 总评成绩按权重自动计算
        - 导出文件名格式：课程名称_班级_学期_成绩单_日期.xlsx
        - 操作记录审计日志
        """

        # ── 核心状态变量（TDD 测试要求） ──
        class_filter: str = ""
        course_filter: str = ""
        is_exporting: bool = False

        # ── 导出列定义 ──
        EXPORT_COLUMNS: List[str] = [
            "学号", "姓名", "班级", "课程名称",
            "出勤得分", "考试得分", "代码提交得分", "PR贡献得分",
            "总评成绩", "等级", "备注",
        ]

        # ── 文件名模板 ──
        FILENAME_TEMPLATE: str = "{course_name}_{class_name}_{term}_成绩单_{date}.xlsx"

        # ── 业务状态变量 ──
        term_filter: str = ""
        preview_rows: List[dict] = []
        selected_weights: dict = {}
        audit_log: List[dict] = []
        filter_options: dict = {}
        export_error: str = ""
        export_message: str = ""

        def get_export_filename(self, course_name: str = "", class_name: str = "", term: str = "") -> str:
            """生成导出文件名：课程名称_班级_学期_成绩单_日期.xlsx"""
            from datetime import datetime
            date_str = datetime.now().strftime("%Y%m%d")
            safe_course = (course_name or self.course_filter or "课程").replace("/", "-")
            safe_class = (class_name or self.class_filter or "全班").replace("/", "-")
            safe_term = (term or self.term_filter or "").replace("/", "-")
            return f"{safe_course}_{safe_class}_{safe_term}_成绩单_{date_str}.xlsx"

        def compute_total_score(self, row: dict, weights: Optional[dict] = None) -> Optional[float]:
            """根据权重公式自动计算总评成绩"""
            return _compute_total_score(row, weights)

        def compute_grade(self, score: float) -> str:
            """根据分数计算等级"""
            return _compute_grade(score)

        async def load_filters(self):
            """加载筛选选项"""
            self.filter_options = _load_filter_options()

        async def preview(self):
            """按当前筛选条件加载预览数据"""
            self.is_exporting = True
            self.export_error = ""
            self.export_message = ""

            try:
                from oaepp.database import db_sync
            except ImportError:
                try:
                    from database import db_sync
                except ImportError:
                    self.export_error = "数据库连接不可用"
                    self.is_exporting = False
                    return

            try:
                with db_sync() as cur:
                    if self.class_filter:
                        cur.execute(
                            """SELECT u.id, u.student_no, u.full_name, s.class_name
                               FROM users u
                               JOIN students s ON u.id = s.user_id
                               WHERE u.role='student' AND s.class_name=%s
                               ORDER BY u.student_no""",
                            (self.class_filter,)
                        )
                    else:
                        cur.execute(
                            """SELECT u.id, u.student_no, u.full_name, s.class_name
                               FROM users u
                               JOIN students s ON u.id = s.user_id
                               WHERE u.role='student'
                               ORDER BY u.student_no"""
                        )
                    students = cur.fetchall()

                    score_map = {}
                    if self.course_filter:
                        cur.execute("SELECT id FROM courses WHERE name=%s", (self.course_filter,))
                        course_row = cur.fetchone()
                        if not course_row:
                            cur.execute("SELECT id FROM courses WHERE id=%s", (self.course_filter,))
                            course_row = cur.fetchone()
                        if course_row:
                            cid = course_row["id"]
                            cur.execute(
                                "SELECT student_user_id, score_type, score FROM score_items WHERE course_id=%s",
                                (cid,)
                            )
                            for r in cur.fetchall():
                                uid = r["student_user_id"]
                                if uid not in score_map:
                                    score_map[uid] = {}
                                score_map[uid][r["score_type"]] = float(r["score"])

                rows = []
                for s in students:
                    uid = s["id"]
                    scores = score_map.get(uid, {})
                    row = {
                        "student_id": s["student_no"],
                        "name": s["full_name"],
                        "class_name": s["class_name"],
                        "course_name": self.course_filter or "",
                        "attendance": scores.get("attendance"),
                        "exam": scores.get("exam"),
                        "code": scores.get("code"),
                        "pr": scores.get("pr"),
                        "total": None,
                        "grade": "",
                        "remark": "",
                    }
                    weights = self.selected_weights or None
                    total = _compute_total_score(row, weights)
                    row["total"] = total
                    row["grade"] = _compute_grade(total)
                    rows.append(row)

                self.preview_rows = rows
                self.export_message = f"预览成功，共 {len(rows)} 条记录"
            except Exception as e:
                logger.error("preview failed: %s", e)
                self.export_error = f"预览失败：{e}"
                self.preview_rows = []
            finally:
                self.is_exporting = False

        async def export_excel(self):
            """生成并下载 Excel，同时写入审计日志"""
            self.is_exporting = True
            self.export_error = ""
            self.export_message = ""

            try:
                from oaepp.database import db_sync, transaction_sync
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                import json
            except ImportError as e:
                self.export_error = f"缺少依赖：{e}"
                self.is_exporting = False
                return

            try:
                import io
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = (self.course_filter or "成绩")[:31]

                header_fill = PatternFill("solid", fgColor="4472C4")
                header_font = Font(bold=True, color="FFFFFF")
                headers = ["学号", "姓名", "班级", "课程名称", "出勤得分", "考试得分", "代码提交得分", "PR贡献得分", "总评成绩", "等级", "备注"]
                widths = [14, 12, 20, 20, 10, 10, 10, 10, 10, 8, 20]

                for col, (h, w) in enumerate(zip(headers, widths), 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                rows = self.preview_rows
                for i, r in enumerate(rows, 2):
                    ws.cell(row=i, column=1, value=r.get("student_id"))
                    ws.cell(row=i, column=2, value=r.get("name"))
                    ws.cell(row=i, column=3, value=r.get("class_name"))
                    ws.cell(row=i, column=4, value=r.get("course_name"))
                    ws.cell(row=i, column=5, value=r.get("attendance"))
                    ws.cell(row=i, column=6, value=r.get("exam"))
                    ws.cell(row=i, column=7, value=r.get("code"))
                    ws.cell(row=i, column=8, value=r.get("pr"))
                    ws.cell(row=i, column=9, value=r.get("total"))
                    ws.cell(row=i, column=10, value=r.get("grade"))
                    ws.cell(row=i, column=11, value=r.get("remark"))

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                filename = self.get_export_filename()

                try:
                    filters = {"class_name": self.class_filter, "course_name": self.course_filter, "term": self.term_filter}
                    with transaction_sync() as cur:
                        cur.execute(
                            "INSERT INTO export_logs (actor, filters, record_count) VALUES (%s, %s, %s)",
                            ("teacher", json.dumps(filters, ensure_ascii=False), len(rows))
                        )
                except Exception as e:
                    logger.warning("Failed to write audit log: %s", e)

                self.is_exporting = False
                return rx.download(data=buf.getvalue(), filename=filename)
            except Exception as e:
                logger.error("export failed: %s", e)
                self.export_error = f"导出失败：{e}"
                self.is_exporting = False

        def add_audit_log(self, actor: str, filters: dict, record_count: int) -> None:
            """添加审计日志记录"""
            from datetime import datetime
            self.audit_log.append({
                "actor": actor,
                "filters": filters,
                "record_count": record_count,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })

        def clear_preview(self) -> None:
            """清空预览数据"""
            self.preview_rows = []
            self.is_exporting = False

        async def load_audit_logs(self):
            """从数据库加载审计日志"""
            self.export_error = ""
            self.export_message = ""

            try:
                from oaepp.database import db_sync
            except ImportError:
                try:
                    from database import db_sync
                except ImportError:
                    logger.warning("db_sync not available, audit logs empty")
                    self.audit_log = []
                    return

            try:
                with db_sync() as cur:
                    cur.execute(
                        "SELECT id, actor, filters, record_count, created_at "
                        "FROM export_logs ORDER BY id DESC LIMIT 200"
                    )
                    self.audit_log = [dict(r) for r in cur.fetchall()]
            except Exception as e:
                logger.error("Failed to load audit logs: %s", e)
                self.audit_log = []
